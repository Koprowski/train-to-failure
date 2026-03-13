#!/usr/bin/env python3
"""
Sync the exercise library from OneFootExerciseList.xlsx.

Workbook columns:
- E: target canonical exercise name
- I: target GIF filename or source asset path
- J: current app/library name, if renaming from an old record
- K: update notes written back by this script on apply

By default the script runs in dry-run mode and prints the planned changes.
Use --apply to write changes. Removals require explicit selection via
--remove all or --remove 1,3,5.
"""

from __future__ import annotations

import argparse
import json
import shutil
import subprocess
import sys
import tempfile
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


REPO_ROOT = Path(__file__).resolve().parent.parent
WORKBOOK_PATH = REPO_ROOT / "OneFootExerciseList.xlsx"
MANIFEST_PATH = REPO_ROOT / "public" / "gifs" / "exercises.json"
GIF_DIR = REPO_ROOT / "public" / "gifs"
DEFAULT_SOURCE_DIR = Path(r"E:\OneDrive\Apps\train-to-failure\gifs")


@dataclass
class WorkbookRow:
    row_num: int
    target_name: str
    source_value: str
    source_path: Path | None
    current_name: str | None
    add_flag: bool
    note: str | None = None


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Sync exercise library from workbook")
    parser.add_argument("--workbook", default=str(WORKBOOK_PATH), help="Path to OneFootExerciseList.xlsx")
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Print the sync plan without writing changes (this is already the default behavior)",
    )
    parser.add_argument("--apply", action="store_true", help="Apply changes instead of printing a dry-run report")
    parser.add_argument(
        "--confirm-remove",
        action="store_true",
        help="Backward-compatible alias for removing all listed candidates when used with --apply",
    )
    parser.add_argument(
        "--remove",
        help="Removal selection for apply mode: 'all' or comma-separated candidate numbers from the dry-run report",
    )
    return parser.parse_args()


def run_command(command: list[str], *, cwd: Path | None = None, input_text: str | None = None) -> str:
    result = subprocess.run(
        command,
        cwd=str(cwd or REPO_ROOT),
        input=input_text,
        capture_output=True,
        text=True,
        check=False,
    )
    if result.returncode != 0:
        raise RuntimeError(
            f"Command failed ({result.returncode}): {' '.join(command)}\n"
            f"STDOUT:\n{result.stdout}\nSTDERR:\n{result.stderr}"
        )
    return result.stdout


def run_node_json(script: str, *, args: list[str] | None = None) -> Any:
    output = run_command(["node", "-", *(args or [])], input_text=script)
    lines = [line for line in output.splitlines() if line.strip()]
    if not lines:
        return None
    return json.loads(lines[0])


def load_seed_exercises() -> list[dict[str, Any]]:
    output = run_command(
        [
            "npx.cmd",
            "tsx",
            "-e",
            (
                "import { SEED_EXERCISES } from './src/lib/seed-exercises.ts'; "
                "console.log(JSON.stringify(SEED_EXERCISES));"
            ),
        ]
    )
    return json.loads(output.splitlines()[-1])


def load_db_state() -> dict[str, Any]:
    script = r"""
const { PrismaClient } = require('@prisma/client');
const prisma = new PrismaClient();

async function main() {
  const official = await prisma.exercise.findMany({
    where: { isCustom: false },
    select: {
      id: true,
      name: true,
      muscleGroups: true,
      equipment: true,
      type: true,
      videoUrl: true,
      imageUrl: true,
      instructions: true,
      links: true,
      _count: {
        select: {
          workoutSets: true,
          workoutExercises: true,
          templateExercises: true,
          favoritedBy: true,
        },
      },
    },
    orderBy: { name: 'asc' },
  });

  const all = await prisma.exercise.findMany({
    select: { id: true, name: true, isCustom: true, imageUrl: true },
    orderBy: { name: 'asc' },
  });

  console.log(JSON.stringify({ official, all }));
}

main()
  .catch((error) => {
    console.error(error);
    process.exit(1);
  })
  .finally(async () => {
    await prisma.$disconnect();
  });
"""
    return run_node_json(script)


def truthy(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    return str(value).strip().lower() in {"true", "1", "yes", "y"}


def normalize_text(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    text = str(value).strip()
    if not text or text == "0":
        return None
    return text


def resolve_source_path(source_value: str, workbook_path: Path) -> Path | None:
    raw = source_value.strip()
    candidates = [
        Path(raw),
        workbook_path.parent / raw,
        workbook_path.parent / "gifs" / raw,
        GIF_DIR / raw,
        DEFAULT_SOURCE_DIR / raw,
    ]

    for candidate in candidates:
        if candidate.exists():
            return candidate
    return None


def load_workbook_rows(workbook_path: Path) -> list[WorkbookRow]:
    workbook = load_workbook(workbook_path, data_only=False)
    sheet = workbook.active
    rows: list[WorkbookRow] = []

    for row_num in range(2, sheet.max_row + 1):
        target_name = normalize_text(sheet[f"E{row_num}"].value)
        source_value = normalize_text(sheet[f"I{row_num}"].value)
        current_name = normalize_text(sheet[f"J{row_num}"].value)
        add_flag = truthy(sheet[f"H{row_num}"].value)

        if not target_name:
            continue

        if not source_value:
            raise RuntimeError(f"Workbook row {row_num} is missing column I (source GIF).")

        rows.append(
            WorkbookRow(
                row_num=row_num,
                target_name=target_name,
                source_value=source_value,
                source_path=resolve_source_path(source_value, workbook_path),
                current_name=current_name,
                add_flag=add_flag,
            )
        )

    return rows


def choose_metadata_source(
    row: WorkbookRow,
    db_by_name: dict[str, dict[str, Any]],
    seed_by_name: dict[str, dict[str, Any]],
) -> tuple[dict[str, Any] | None, str | None]:
    candidates = [row.current_name, row.target_name]
    for candidate in candidates:
        if candidate and candidate in db_by_name:
            return db_by_name[candidate], f"db:{candidate}"
    for candidate in candidates:
        if candidate and candidate in seed_by_name:
            return seed_by_name[candidate], f"seed:{candidate}"
    return None, None


def build_sync_plan(
    workbook_rows: list[WorkbookRow],
    db_state: dict[str, Any],
    seed_exercises: list[dict[str, Any]],
) -> dict[str, Any]:
    official_rows = db_state["official"]
    all_rows = db_state["all"]
    db_by_name = {row["name"]: row for row in official_rows}
    seed_by_name = {row["name"]: row for row in seed_exercises}
    target_names = {row.target_name for row in workbook_rows}

    desired_records: list[dict[str, Any]] = []
    row_notes: dict[int, str] = {}
    warnings: list[str] = []
    db_upserts: list[dict[str, Any]] = []

    for row in workbook_rows:
        metadata, metadata_source = choose_metadata_source(row, db_by_name, seed_by_name)
        if metadata is None:
            warning = (
                f"Row {row.row_num}: no metadata source found for target '{row.target_name}' "
                f"(current='{row.current_name or ''}')"
            )
            warnings.append(warning)
            row_notes[row.row_num] = "Blocked: missing metadata source"
            continue

        if row.source_path is None:
            warning = (
                f"Row {row.row_num}: source GIF not found for '{row.target_name}' "
                f"({row.source_value})"
            )
            warnings.append(warning)
            row_notes[row.row_num] = "Blocked: source GIF not found"
            continue

        gif_file = row.source_path.name
        image_url = f"/gifs/{gif_file}"
        desired = {
            "name": row.target_name,
            "muscleGroups": metadata["muscleGroups"],
            "equipment": metadata["equipment"],
            "type": metadata["type"],
            "videoUrl": metadata.get("videoUrl"),
            "instructions": metadata.get("instructions"),
            "links": metadata.get("links"),
            "gifFile": gif_file,
            "imageUrl": image_url,
            "sourcePath": str(row.source_path),
            "rowNum": row.row_num,
        }
        desired_records.append(desired)

        target_row = db_by_name.get(row.target_name)
        current_row = db_by_name.get(row.current_name) if row.current_name else None

        update_source = current_row or target_row
        if target_row:
            action = "No change"
            changes: list[str] = []
            for field in ("muscleGroups", "equipment", "type", "videoUrl", "instructions", "links", "imageUrl"):
                if (target_row.get(field) or None) != (desired.get(field) or None):
                    changes.append(field)
            if changes:
                action = f"Update {', '.join(changes)}"
                db_upserts.append({"mode": "update", "id": target_row["id"], "data": desired})
            if current_row and current_row["id"] != target_row["id"]:
                action = f"{action}; current duplicate '{row.current_name}' will be flagged for removal"
            row_notes[row.row_num] = action
        elif current_row:
            db_upserts.append({"mode": "update", "id": current_row["id"], "data": desired})
            row_notes[row.row_num] = f"Rename from {row.current_name}"
        else:
            db_upserts.append({"mode": "create", "data": desired})
            if row.add_flag:
                row_notes[row.row_num] = "Add new exercise"
            else:
                row_notes[row.row_num] = f"Create missing exercise ({metadata_source})"

        if update_source and row.current_name and row.current_name == row.target_name and row.add_flag:
            row_notes[row.row_num] = "Exists already; add flag can be cleared"

    removal_candidates: list[dict[str, Any]] = []
    desired_image_urls = {record["imageUrl"] for record in desired_records}
    remaining_non_removed_images = {
        row["imageUrl"]
        for row in all_rows
        if row["name"] in target_names or row["isCustom"]
    }

    for row in official_rows:
        if row["name"] in target_names:
            continue
        counts = row.get("_count", {})
        total_refs = sum(int(counts.get(key, 0)) for key in counts)
        image_url = row.get("imageUrl")
        image_file = Path(image_url).name if image_url else None
        orphan_gif = bool(
            image_url
            and image_url not in desired_image_urls
            and image_url not in remaining_non_removed_images
        )
        removal_candidates.append(
            {
                "id": row["id"],
                "name": row["name"],
                "imageUrl": image_url,
                "imageFile": image_file,
                "referenceCounts": counts,
                "totalReferences": total_refs,
                "orphanGif": orphan_gif,
            }
        )

    return {
        "desiredRecords": desired_records,
        "rowNotes": row_notes,
        "warnings": warnings,
        "dbUpserts": db_upserts,
        "removalCandidates": removal_candidates,
    }


def resolve_removal_selection(plan: dict[str, Any], remove_arg: str | None, confirm_remove: bool) -> list[dict[str, Any]]:
    removal_candidates = plan["removalCandidates"]

    if confirm_remove and not remove_arg:
        remove_arg = "all"

    if not remove_arg:
        return []

    if remove_arg.strip().lower() == "all":
        return removal_candidates

    selected_indexes: set[int] = set()
    for part in remove_arg.split(","):
        token = part.strip()
        if not token:
            continue
        if not token.isdigit():
            raise RuntimeError(f"Invalid removal selection '{token}'. Use numbers like 1,3,5 or 'all'.")
        selected_indexes.add(int(token))

    if not selected_indexes:
        return []

    selected: list[dict[str, Any]] = []
    for idx, candidate in enumerate(removal_candidates, start=1):
        if idx in selected_indexes:
            selected.append(candidate)

    unknown = sorted(index for index in selected_indexes if index < 1 or index > len(removal_candidates))
    if unknown:
        raise RuntimeError(f"Removal selection out of range: {', '.join(str(x) for x in unknown)}")

    return selected


def write_manifest(desired_records: list[dict[str, Any]]) -> None:
    manifest: dict[str, Any] = {}
    for record in desired_records:
        manifest[record["name"]] = {
            "gifFile": record["gifFile"],
            "imageUrl": record["imageUrl"],
            "muscleGroups": record["muscleGroups"],
        }
    MANIFEST_PATH.write_text(json.dumps(manifest, indent=2) + "\n", encoding="utf-8")


def format_ts_value(value: Any) -> str:
    if value is None:
        return "undefined"
    return json.dumps(value, ensure_ascii=False)


def build_seed_exercises_source(desired_records: list[dict[str, Any]]) -> str:
    lines = [
        "export interface SeedExercise {",
        "  name: string;",
        "  muscleGroups: string;",
        "  equipment: string;",
        "  type: string;",
        "  videoUrl?: string;",
        "}",
        "",
        "export const SEED_EXERCISES: SeedExercise[] = [",
    ]
    for record in desired_records:
        parts = [
            f'name: {format_ts_value(record["name"])}',
            f'muscleGroups: {format_ts_value(record["muscleGroups"])}',
            f'equipment: {format_ts_value(record["equipment"])}',
            f'type: {format_ts_value(record["type"])}',
        ]
        if record.get("videoUrl"):
            parts.append(f'videoUrl: {format_ts_value(record["videoUrl"])}')
        lines.append(f"  {{ {', '.join(parts)} }},")
    lines.extend(["];", ""])
    return "\n".join(lines)


def build_prisma_seed_source(desired_records: list[dict[str, Any]]) -> str:
    lines = [
        'import { PrismaClient } from "@prisma/client";',
        "",
        "const prisma = new PrismaClient();",
        "",
        "const exercises = [",
    ]
    for record in desired_records:
        parts = [
            f'name: {format_ts_value(record["name"])}',
            f'muscleGroups: {format_ts_value(record["muscleGroups"])}',
            f'equipment: {format_ts_value(record["equipment"])}',
            f'type: {format_ts_value(record["type"])}',
        ]
        if record.get("videoUrl"):
            parts.append(f'videoUrl: {format_ts_value(record["videoUrl"])}')
        lines.append(f"  {{ {', '.join(parts)} }},")
    lines.extend(
        [
            "];",
            "",
            "async function main() {",
            "  await prisma.exercise.deleteMany({ where: { isCustom: false } });",
            "",
            '  console.log("Seeding exercises...");',
            "  for (const exercise of exercises) {",
            "    await prisma.exercise.create({ data: { ...exercise, isCustom: false } });",
            "  }",
            "  console.log(`Seeded ${exercises.length} exercises`);",
            "}",
            "",
            "main()",
            "  .catch((error) => {",
            "    console.error(error);",
            "    process.exit(1);",
            "  })",
            "  .finally(async () => {",
            "    await prisma.$disconnect();",
            "  });",
            "",
        ]
    )
    return "\n".join(lines)


def copy_gifs(desired_records: list[dict[str, Any]]) -> int:
    copied = 0
    for record in desired_records:
        source = Path(record["sourcePath"])
        destination = GIF_DIR / record["gifFile"]
        if not destination.exists() or source.read_bytes() != destination.read_bytes():
            shutil.copy2(source, destination)
            copied += 1
    return copied


def write_workbook_notes(workbook_path: Path, row_notes: dict[int, str]) -> None:
    workbook = load_workbook(workbook_path)
    sheet = workbook.active
    for row_num, note in row_notes.items():
        sheet[f"K{row_num}"] = note
    workbook.save(workbook_path)


def apply_db_changes(plan: dict[str, Any], removals_to_apply: list[dict[str, Any]]) -> dict[str, Any]:
    delete_ids = [
        candidate["id"] for candidate in removals_to_apply if candidate["totalReferences"] == 0
    ]
    payload = {
        "upserts": plan["dbUpserts"],
        "deleteIds": delete_ids,
    }

    with tempfile.NamedTemporaryFile("w", suffix=".json", delete=False, encoding="utf-8") as handle:
        json.dump(payload, handle)
        temp_path = Path(handle.name)

    script = r"""
const fs = require('fs');
const { PrismaClient } = require('@prisma/client');
const prisma = new PrismaClient();
const payload = JSON.parse(fs.readFileSync(process.argv[2], 'utf8'));

async function main() {
  for (const item of payload.upserts) {
    const data = {
      name: item.data.name,
      muscleGroups: item.data.muscleGroups,
      equipment: item.data.equipment,
      type: item.data.type,
      videoUrl: item.data.videoUrl ?? null,
      instructions: item.data.instructions ?? null,
      links: item.data.links ?? null,
      imageUrl: item.data.imageUrl,
      isCustom: false,
      userId: null,
    };

    if (item.mode === 'update') {
      await prisma.exercise.update({ where: { id: item.id }, data });
    } else if (item.mode === 'create') {
      await prisma.exercise.create({ data });
    }
  }

  if (payload.deleteIds.length > 0) {
    await prisma.exercise.deleteMany({
      where: {
        id: { in: payload.deleteIds },
        isCustom: false,
      },
    });
  }

  console.log(JSON.stringify({ updatedOrCreated: payload.upserts.length, deleted: payload.deleteIds.length }));
}

main()
  .catch((error) => {
    console.error(error);
    process.exit(1);
  })
  .finally(async () => {
    await prisma.$disconnect();
  });
"""

    try:
        return run_node_json(script, args=[str(temp_path)])
    finally:
        temp_path.unlink(missing_ok=True)


def delete_orphan_gifs(removal_candidates: list[dict[str, Any]]) -> list[str]:
    deleted: list[str] = []
    for candidate in removal_candidates:
        if not candidate["orphanGif"] or candidate["totalReferences"] != 0 or not candidate["imageFile"]:
            continue
        path = GIF_DIR / candidate["imageFile"]
        if path.exists():
            path.unlink()
            deleted.append(candidate["imageFile"])
    return deleted


def print_report(plan: dict[str, Any], apply: bool, removals_to_apply: list[dict[str, Any]]) -> None:
    mode = "APPLY" if apply else "DRY RUN"
    print(f"{mode}: exercise workbook sync")
    print(f"Desired workbook rows: {len(plan['desiredRecords'])}")
    print(f"DB create/update actions: {len(plan['dbUpserts'])}")

    if plan["warnings"]:
        print("\nWarnings:")
        for warning in plan["warnings"]:
            print(f"  - {warning}")

    if plan["removalCandidates"]:
        print("\nRemoval candidates (not on workbook):")
        selected_ids = {candidate["id"] for candidate in removals_to_apply}
        for idx, candidate in enumerate(plan["removalCandidates"], start=1):
            refs = candidate["totalReferences"]
            orphan = "yes" if candidate["orphanGif"] else "no"
            marker = "*" if candidate["id"] in selected_ids else " "
            print(
                f" {marker} {idx}. {candidate['name']} | refs={refs} | "
                f"orphanGif={orphan} | image={candidate['imageFile'] or 'n/a'}"
            )
    else:
        print("\nRemoval candidates: none")

    if removals_to_apply:
        selected_labels = ", ".join(candidate["name"] for candidate in removals_to_apply)
        print(f"\nSelected for removal on apply: {selected_labels}")
    else:
        print("\nNo removals selected. Extra official records will only be reported.")


def main() -> int:
    args = parse_args()
    workbook_path = Path(args.workbook).resolve()
    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")

    workbook_rows = load_workbook_rows(workbook_path)
    db_state = load_db_state()
    seed_exercises = load_seed_exercises()
    plan = build_sync_plan(workbook_rows, db_state, seed_exercises)
    removals_to_apply = resolve_removal_selection(plan, args.remove, args.confirm_remove)

    if removals_to_apply and not args.apply:
        raise RuntimeError("Removal selection requires --apply.")

    print_report(plan, apply=args.apply, removals_to_apply=removals_to_apply)

    if not args.apply:
        return 0

    copied = copy_gifs(plan["desiredRecords"])
    write_manifest(plan["desiredRecords"])
    (REPO_ROOT / "src" / "lib" / "seed-exercises.ts").write_text(
        build_seed_exercises_source(plan["desiredRecords"]),
        encoding="utf-8",
    )
    (REPO_ROOT / "prisma" / "seed.ts").write_text(
        build_prisma_seed_source(plan["desiredRecords"]),
        encoding="utf-8",
    )
    db_result = apply_db_changes(plan, removals_to_apply=removals_to_apply)
    deleted_gifs = delete_orphan_gifs(removals_to_apply) if removals_to_apply else []

    try:
        write_workbook_notes(workbook_path, plan["rowNotes"])
        workbook_note_status = "Workbook notes updated"
    except PermissionError:
        workbook_note_status = "Workbook notes skipped (file is open/locked)"

    print("\nApplied changes:")
    print(f"  - GIFs copied or refreshed: {copied}")
    print(f"  - DB rows created/updated: {db_result['updatedOrCreated']}")
    print(f"  - DB rows deleted: {db_result['deleted']}")
    print(f"  - GIFs deleted: {len(deleted_gifs)}")
    print(f"  - {workbook_note_status}")

    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as error:
        print(f"ERROR: {error}", file=sys.stderr)
        raise
